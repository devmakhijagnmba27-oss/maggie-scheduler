"""
Maggie Calendar Scheduler – Timetable & Elective Parser
========================================================
Handles:
  1. IILM MBA Timetable Excel files (.xlsx / .xls) with multi-sheet sections (Sec A-G, MT)
  2. PDF timetables
  3. Elective Submission PDFs

Extracts:
  - Dates, Days, Time Slots
  - Course Acronyms in each slot
  - Course Catalog (Acronym -> Course Title, Faculty Name, Credits)
  - Room Allocations (e.g. Room 302, Room 303, Seminar Hall)
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any
import openpyxl
import pandas as pd
import pdfplumber


def parse_timetable_excel(filepath: str | Path, section: str = "Sec D") -> dict:
    """
    Parse an Excel timetable file with high precision for section sheets.
    """
    filepath = Path(filepath)
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # Choose target sheet
    target_sheet_name = None
    # Look for matching section
    for name in wb.sheetnames:
        if section.lower().replace(" ", "") in name.lower().replace(" ", ""):
            target_sheet_name = name
            break
    if not target_sheet_name:
        target_sheet_name = wb.sheetnames[0]  # Fallback to first sheet
        
    sheet = wb[target_sheet_name]
    
    slots: list[dict] = []
    course_map: dict[str, dict] = {}
    room_map: dict[str, str] = {}
    
    # 1. Scan for default room info in header
    for r in range(1, 6):
        for c in range(1, 10):
            val = str(sheet.cell(row=r, column=c).value or "").strip()
            if "room" in val.lower():
                room_map.setdefault("_DEFAULT_ROOM", "302/303")
                
    # 2. Extract Course Catalog & Faculty from the side columns
    # Find row with header 'Acronym' and 'Course Title' and 'Faculty'
    acronym_col = None
    title_col = None
    faculty_col = None
    catalog_start_row = None
    
    for r in range(1, 15):
        for c in range(1, sheet.max_column + 1):
            val = str(sheet.cell(row=r, column=c).value or "").strip().lower()
            if val == "acronym":
                acronym_col = c
                catalog_start_row = r + 1
            elif "course title" in val or val == "course":
                title_col = c
            elif "faculty" in val or "teacher" in val:
                faculty_col = c
                
    if acronym_col and catalog_start_row:
        for r in range(catalog_start_row, sheet.max_row + 1):
            code = str(sheet.cell(row=r, column=acronym_col).value or "").strip().upper()
            if not code or code in ("NONE", "NAN", ""):
                continue
            
            title = str(sheet.cell(row=r, column=title_col).value or "").strip() if title_col else ""
            faculty = str(sheet.cell(row=r, column=faculty_col).value or "").strip() if faculty_col else ""
            
            # Clean up title (remove trailing numbers like - 66, - 134 if present)
            clean_title = re.sub(r'\s*-\s*\d+$', '', title).strip()
            
            course_map[code] = {
                "name": clean_title or title,
                "faculty": faculty if faculty.lower() not in ("none", "nan", "") else "",
            }

    # 3. Extract room notes anywhere in the sheet (e.g. 'B2B, CB, WM & PBM in 302. CM in 303.')
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            val = str(sheet.cell(row=r, column=c).value or "").strip()
            if " in " in val.lower() and any(d in val for d in ["103", "205", "302", "303", "Seminar", "VOICE"]):
                # Parse sentences like 'B2B, CB, WM & PBM in 302'
                clauses = re.split(r'[.;\n]+', val)
                for clause in clauses:
                    m = re.match(r'(.+?)\s+in\s+([A-Za-z0-9/\s]+)', clause.strip(), re.IGNORECASE)
                    if m:
                        codes_part = m.group(1).replace("&", ",").replace("/", ",")
                        room_name = m.group(2).strip()
                        for code_token in codes_part.split(","):
                            code_token = code_token.strip().upper()
                            if code_token and len(code_token) <= 8:
                                room_map[code_token] = room_name

    # 4. Extract Timetable Grid
    # Header row with time slots
    time_cols: dict[int, str] = {}
    time_row = None
    date_col = 1
    day_col = 2
    
    for r in range(1, 10):
        for c in range(1, 15):
            val = str(sheet.cell(row=r, column=c).value or "").strip()
            # Match 9:00 - 10:15 or 10:20-11:35
            if re.search(r'\d{1,2}[:.]\d{2}\s*[-–]\s*\d{1,2}[:.]\d{2}', val):
                time_cols[c] = val.replace(" ", "")
                time_row = r
                
    if time_row:
        # Collect all rows with day info
        raw_day_rows = []
        for r in range(time_row + 1, sheet.max_row + 1):
            date_val = str(sheet.cell(row=r, column=date_col).value or "").strip()
            day_val = str(sheet.cell(row=r, column=day_col).value or "").strip()
            day_name = _normalize_day(day_val)
            if not day_name:
                continue
            raw_day_rows.append((r, date_val, day_name))

        # Check if the sheet contains multiple weeks (e.g. 03-Aug, 10-Aug, 17-Aug)
        # We want the active latest week block (e.g. the last week or WEF week)
        active_day_rows = []
        if len(raw_day_rows) > 7:
            # Group into weeks by detecting Monday restarts
            weeks = []
            current_week = []
            for item in raw_day_rows:
                if item[2] == "Monday" and current_week:
                    weeks.append(current_week)
                    current_week = [item]
                else:
                    current_week.append(item)
            if current_week:
                weeks.append(current_week)
            
            # Select the last full week block (the WEF active week)
            active_day_rows = weeks[-1]
        else:
            active_day_rows = raw_day_rows

        for r, date_val, day_name in active_day_rows:
            # Date formatting (e.g. '2026-08-17 00:00:00' -> '17-Aug')
            date_str = ""
            if date_val:
                m_date = re.search(r'(\d{4})-(\d{2})-(\d{2})', date_val)
                if m_date:
                    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
                    m_idx = int(m_date.group(2)) - 1
                    date_str = f"{m_date.group(3)}-{month_names[m_idx]}"
                else:
                    date_str = date_val.split()[0]

            for c, time_str in time_cols.items():
                cell_val = str(sheet.cell(row=r, column=c).value or "").strip()
                if not cell_val or cell_val.lower() in ("nan", "-", "none", ""):
                    continue
                    
                entries = _split_cell_entries(cell_val)
                if entries:
                    slots.append({
                        "date": date_str,
                        "day": day_name,
                        "time": time_str,
                        "entries": entries,
                        "raw_cell": cell_val,
                    })

    return {"slots": slots, "course_map": course_map, "room_map": room_map}


def parse_timetable_pdf(filepath: str | Path) -> dict:
    """
    Parse a PDF timetable by extracting tables and text.
    """
    filepath = Path(filepath)
    all_tables: list[list[list[str]]] = []
    full_text = ""

    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for t in tables:
                cleaned = [[str(cell).strip() if cell else "" for cell in row] for row in t]
                all_tables.append(cleaned)
            text = page.extract_text() or ""
            full_text += text + "\n"

    slots: list[dict] = []
    course_map: dict[str, dict] = {}
    room_map: dict[str, str] = {}

    for table in all_tables:
        df = pd.DataFrame(table).fillna("").astype(str)
        day_row_idx, time_cols, day_col_idx = _find_grid_axes(df)
        if day_row_idx is not None:
            slots.extend(_extract_grid_slots(df, day_row_idx, time_cols, day_col_idx))
        cm = _extract_course_map(df)
        if cm:
            course_map.update(cm)
        rm = _extract_room_map(df)
        if rm:
            room_map.update(rm)

    cm_text = _extract_course_map_from_text(full_text)
    if cm_text:
        course_map.update(cm_text)

    rm_text = _extract_room_map_from_text(full_text)
    if rm_text:
        room_map.update(rm_text)

    return {"slots": slots, "course_map": course_map, "room_map": room_map}


def parse_electives_pdf(filepath: str | Path) -> list[str]:
    """
    Extract chosen elective subject acronyms from the student's submission PDF.
    Maps course names like 'B2B Marketing' -> 'B2B', 'Data Driven Marketing' -> 'DDM', etc.
    """
    filepath = Path(filepath)
    text = ""
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            text += (page.extract_text() or "") + "\n"

    # Well-known mapping
    name_to_acronym = {
        "B2B MARKETING": "B2B",
        "DIGITAL & SOCIAL MEDIA MARKETING": "DSMM",
        "CONSUMER BEHAVIOUR": "CB",
        "CONSUMER BEHAVIOR": "CB",
        "PRODUCT AND BRAND MANAGEMENT": "PBM",
        "DATA DRIVEN MARKETING": "DDM",
        "EMOTIONAL INTELLIGENCE": "EI",
        "CORPORATE READINESS PROGRAM": "CRP3",
        "CORPORATE READINESS PROGRAM-3": "CRP3",
        "FINTECH": "FINTECH",
        "DATA ANALYTICS": "DA",
    }

    subjects: list[str] = []
    for line in text.splitlines():
        line_upper = line.strip().upper()
        for full_title, acronym in name_to_acronym.items():
            if full_title in line_upper and acronym not in subjects:
                subjects.append(acronym)

    # Fallback to direct codes if none matched
    if not subjects:
        for full_title, acronym in name_to_acronym.items():
            subjects.append(acronym)

    return subjects


def parse_timetable(filepath: str | Path, section: str = "Sec D") -> dict:
    """Auto-detect format and parse timetable."""
    filepath = Path(filepath)
    ext = filepath.suffix.lower()
    if ext in (".xlsx", ".xls"):
        return parse_timetable_excel(filepath, section=section)
    elif ext == ".pdf":
        return parse_timetable_pdf(filepath)
    else:
        raise ValueError(f"Unsupported timetable format: {ext}")


# ── Helpers ──────────────────────────────────────────────

def _normalize_day(raw: str) -> str | None:
    raw = raw.strip().upper()
    day_map = {
        "MON": "Monday", "MONDAY": "Monday",
        "TUE": "Tuesday", "TUESDAY": "Tuesday",
        "WED": "Wednesday", "WEDNESDAY": "Wednesday",
        "THU": "Thursday", "THURSDAY": "Thursday",
        "FRI": "Friday", "FRIDAY": "Friday",
        "SAT": "Saturday", "SATURDAY": "Saturday",
        "SUN": "Sunday", "SUNDAY": "Sunday",
    }
    return day_map.get(raw)


def _split_cell_entries(cell: str) -> list[str]:
    cell = cell.replace("\n", "/").replace("\\n", "/").replace(",", "/")
    parts = re.split(r'[/|]+', cell)
    results = []
    for p in parts:
        p = p.strip().upper()
        p = re.sub(r'[^A-Z0-9&\s]', '', p).strip()
        if p and len(p) >= 2:
            results.append(p)
    return results


def _find_grid_axes(df: pd.DataFrame) -> tuple:
    header_row_idx = None
    time_cols: dict[int, str] = {}
    day_col_idx = None
    time_pattern = re.compile(r'\d{1,2}[:.]\d{2}\s*[-–]\s*\d{1,2}[:.]\d{2}')

    best_row = None
    best_count = 0
    for idx, row in df.iterrows():
        count = sum(1 for cell in row if time_pattern.search(str(cell)))
        if count > best_count:
            best_count = count
            best_row = idx
    if best_count >= 2:
        header_row_idx = best_row
        for col_idx, cell in enumerate(df.iloc[header_row_idx]):
            m = time_pattern.search(str(cell))
            if m:
                time_cols[col_idx] = m.group().strip().replace(" ", "")

    day_names = {"MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "MON", "TUE", "WED", "THU", "FRI", "SAT"}
    for col_idx in range(min(3, df.shape[1])):
        day_count = 0
        for idx in range(df.shape[0]):
            cell = str(df.iloc[idx, col_idx]).strip().upper()
            for token in cell.replace("\n", " ").split():
                if token in day_names:
                    day_count += 1
                    break
        if day_count >= 3:
            day_col_idx = col_idx
            break

    return header_row_idx, time_cols, day_col_idx


def _extract_grid_slots(df: pd.DataFrame, header_row: int, time_cols: dict[int, str], day_col: int | None) -> list[dict]:
    slots = []
    for row_idx in range(header_row + 1, df.shape[0]):
        day_raw = str(df.iloc[row_idx, day_col]).strip() if day_col is not None else ""
        day_name = None
        for token in day_raw.replace("\n", " ").upper().split():
            d = _normalize_day(token)
            if d:
                day_name = d
                break
        if not day_name:
            continue

        for col_idx, time_str in time_cols.items():
            cell = str(df.iloc[row_idx, col_idx]).strip()
            if not cell or cell.lower() in ("nan", "-", ""):
                continue
            entries = _split_cell_entries(cell)
            if entries:
                slots.append({
                    "date": "",
                    "day": day_name,
                    "time": time_str,
                    "entries": entries,
                    "raw_cell": cell,
                })
    return slots


def _extract_course_map(df: pd.DataFrame) -> dict[str, dict]:
    course_map = {}
    for idx, row in df.iterrows():
        row_str = " ".join(str(c).upper() for c in row)
        if "ACRONYM" in row_str:
            headers = [str(c).strip().upper() for c in row]
            code_col = _find_col_idx(headers, ["ACRONYM", "CODE"])
            name_col = _find_col_idx(headers, ["COURSE TITLE", "COURSE NAME", "TITLE"])
            fac_col = _find_col_idx(headers, ["FACULTY", "TEACHER", "PROF"])
            for r in range(idx + 1, df.shape[0]):
                code = str(df.iloc[r, code_col]).strip().upper() if code_col is not None else ""
                if code and code not in ("NAN", "NONE", ""):
                    name = str(df.iloc[r, name_col]).strip() if name_col is not None else ""
                    fac = str(df.iloc[r, fac_col]).strip() if fac_col is not None else ""
                    course_map[code] = {
                        "name": re.sub(r'\s*-\s*\d+$', '', name).strip(),
                        "faculty": fac if fac not in ("NAN", "NONE") else "",
                    }
            break
    return course_map


def _extract_course_map_from_text(text: str) -> dict[str, dict]:
    course_map = {}
    for line in text.splitlines():
        m = re.match(r'^([A-Z0-9&]{2,10})\s+([A-Za-z\s&–-]+?)\s+(?:2|3|4)?\s*(?:\d+)?\s*(?:\d+)?\s*([A-Za-z\s.]+)?$', line.strip())
        if m:
            code = m.group(1).strip().upper()
            name = m.group(2).strip()
            faculty = m.group(3).strip() if m.group(3) else ""
            if len(code) <= 8 and len(name) > 3:
                course_map[code] = {"name": name, "faculty": faculty}
    return course_map


def _extract_room_map(df: pd.DataFrame) -> dict[str, str]:
    return {}


def _extract_room_map_from_text(text: str) -> dict[str, str]:
    room_map = {}
    pattern = re.compile(r'((?:[A-Z0-9&]+(?:\s*[,&]\s*)?)+)\s+in\s+(\d+\w*(?:\s*/\s*\d+\w*)*)', re.IGNORECASE)
    for m in pattern.finditer(text):
        codes = re.split(r'[,&]+', m.group(1))
        room = m.group(2).strip()
        for c in codes:
            c = c.strip().upper()
            if c:
                room_map[c] = room
    return room_map


def _find_col_idx(headers: list[str], keywords: list[str]) -> int | None:
    for i, h in enumerate(headers):
        for kw in keywords:
            if kw in h:
                return i
    return None
